from openpyxl import load_workbook
wb = load_workbook('sample.xlsx')
ws = wb.active


# print(english_col)
for row in ws.iter_rows(min_row = 2):
    if row[1].value > 80:
        print(f'{row[0].value}번 학생은 영어 천재입니다.')
    
for row in ws.iter_rows(max_row = 1):
    if row[1].value == '영어':
        row[1].value = '컴퓨터'

wb.save('sample_modify.xlsx')
wb.close()
