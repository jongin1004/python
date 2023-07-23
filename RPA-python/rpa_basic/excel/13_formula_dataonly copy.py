import datetime
from openpyxl import load_workbook
wb = load_workbook('sample_modify.xlsx')
ws = wb.active

# 이렇게 바로 값을 출력하면, 수식이 그대로 TEXT로 출력이됨
for row in ws.values:
    for cell in row:
        print(cell)


wb = load_workbook('sample_modify.xlsx', data_only = True)
ws = wb.active

# evaluate 계산되지 않은 상태의 데이터가 None
# excel 파일을 열어서 저장한 뒤에, 종료하게 되면 수식이 계산된 값으로써 저장되기 때문에 None이 아님
for row in ws.values:
    for cell in row:
        print(cell)


# wb.save('sample_modify.xlsx')
# wb.close()
