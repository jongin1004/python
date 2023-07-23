from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl import load_workbook
wb = load_workbook('sample.xlsx')
ws = wb.active

# 번호, 영어, 수학
a1 = ws['A1']
b1 = ws['B1']
c1 = ws['C1']

# A 열의 너비를 4로설정
ws.column_dimensions['A'].width = 4
# 1 행의 높이를 50으로 설정
ws.row_dimensions[1].height = 50

a1.font = Font(color = 'FF0000', italic = True, bold = True)
# name 폰트를 지정, strike는 글자에 가운데 선을 그음
b1.font = Font(color = 'CC33FF', name = 'Arial', strike = True)
# size: font의 사이즈, underline: 글자 밑에 밑줄을 긋는 
c1.font = Font(color = '0000FF', size = 20, underline = 'single')

# 테두리 적용
thin_border = Border(left = Side(style = 'thin'), right = Side(style = 'thin'), top = Side(style = 'thin'), bottom = Side(style = 'thin'))
a1.border = thin_border 
b1.border = thin_border 
c1.border = thin_border 

# 점수가 90점 이상인 cell의 색 변경 
for row in ws.rows:
    for cell in row:
        # 가운데 정렬 (center/left/right/top/bottom)
        cell.alignment = Alignment(horizontal = 'center', vertical = 'center')
        if cell.column == 1: # A 열 제외
            continue

        if isinstance(cell.value, str):
            continue

        if cell.value >= 90:
            cell.fill = PatternFill(fgColor='00FF00', fill_type = 'solid')
            cell.font = Font(color = 'FFFFFF')

# 툴고정 (행고정)
ws.freeze_panes = 'B2' #B2를 기준으로 그 위의 행과, 그 왼쪽의 열이 고정이됨
wb.save('sample_modify.xlsx')
wb.close()
