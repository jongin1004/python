from openpyxl import Workbook
from random import *
wb = Workbook()
ws = wb.active

ws.append(['번호', '영어', '수학']) # 1줄씩 데이터 넣기
for i in range(1, 11):
    ws.append([i, randint(0, 100), randint(0, 100)])

col_b = ws['B'] # b의 column 가져오기 
# for cell in col_b:
#     print(cell.value)

col_range = ws["B:C"] # b~c 컬럼을 모두 획득
for cols in col_range:
    for cell in cols:
        print(cell.value, end = " ")
    print()

row_title = ws[1] # 첫 번째 row만 가져오기
# print(row_title)
# for cell in row_title:
#     print(cell.value)

row_range = ws[2:6] # 2번째 줄부터 6번째 줄까지 포함해서 가져오기
# for rows in row_range:
#     for cell in rows:
#         print(cell.value, end = " ")
#     print()

from openpyxl.utils.cell import coordinate_from_string

row_range = ws[2:ws.max_row] # 2번째 줄부터 마지막 줄까지
# for rows in row_range:
#     for cell in rows:
        # print(cell.value, end = " ")
        # print(cell.coordinate, end = " ") # cell.coordinate 셀의 좌표값을 표시 (A1, B1 ...)
        # xy = coordinate_from_string(cell.coordinate) # 좌표값을 문자와 숫자를 분리해서 튜플형태로 (A1 => ('A', 1))
        # print(xy, end = " ")

    #     print(xy[0], end = "")
    #     print(xy[1], end = " ")
    # print()

# 전체 rows 가져오기 
# print(tuple(ws.rows))
# for row in tuple(ws.rows):
#     print(row)
# for row in ws.iter_rows():
#     print(row)

# 전체 column 가져오기
# print(tuple(ws.columns))
# for row in tuple(ws.columns):
#     print(row)
# for col in ws.iter_cols():
#     print(col)

# iter_cols와 iter_rows 는 max와 min을 설정할 수 있다.
for row in ws.iter_rows(min_row = 1, max_row = 5, min_col = 2, max_col = 3):
    # print(row)
    print(f"{'수학:'}{row[0].value}", f"{'영어:'}{row[1].value}")

for col in ws.iter_cols(min_row = 1, max_row = 5, min_col = 2, max_col = 3):
    print(col)
    # print(f"{'수학:'}{row[0].value}", f"{'영어:'}{row[1].value}")

wb.save('sample.xlsx') # 파일 저장
wb.close() # 파일 종료 
