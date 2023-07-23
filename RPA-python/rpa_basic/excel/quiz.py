from openpyxl import load_workbook

wb = load_workbook('quiz.xlsx')
ws = wb.active

ws['H1'] = '총점'
ws['I1'] = '성적'
for idx, rows in enumerate(ws.rows, start = 1):
    if idx == 1:
        continue

    sum_value = 0
    grade     = 'F'
    # 퀴즈 2를 모두 만점 처리 (10점)
    for cell in rows:
        if cell.column == 4:
            cell.value = 10

        if cell.column > 1 and cell.column <= 7:
            sum_value += int(cell.value)                                        
        
    # H열에 총점(SUM 서식)
    ws.cell(row = idx, column = 8).value = f'=SUM(B{idx}:G{idx})'        
      
    # 성적  
    if sum_value >= 90:
        grade = 'A' 
    elif sum_value >= 80:
        grade = 'B' 
    elif sum_value >= 70:
        grade = 'C' 
    else :
        grade = 'D' 

    # 출석 체크
    if ws.cell(row = idx, column = 2).value < 5:
        grade = 'F'

    ws.cell(row = idx, column = 9).value = grade

wb.save('quiz_result.xlsx')
wb.close()