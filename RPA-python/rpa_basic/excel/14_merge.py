import datetime
from openpyxl import Workbook
wb = Workbook()
ws = wb.active

ws.merge_cells('B2:D2') # B2~D2 병합
ws['B2'].value = 'Merged Cell'

ws.unmerge_cells('B2:D2') # 병합해제

wb.save('sample_modify.xlsx')
wb.close()
