from openpyxl import load_workbook
wb = load_workbook('sample.xlsx')
ws = wb.active

# 번호, 영어, 수학 -> 번호, 국어, 영어, 수학
# 이동하려는 범위의 값을 같은행(rows=0)에 열만 1칸 오른쪽(cols=1)으로 이동하도록
# 값을 
# ws.move_range("B1:C11", rows=0, cols=1) 
# ws['B1'].value = '국어'

# -는 반대 방향으로 이동 
ws.move_range("C1:C11", rows=5, cols=-1) 

wb.save('sample_modify.xlsx')
wb.close()
