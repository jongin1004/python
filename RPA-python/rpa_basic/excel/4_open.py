from openpyxl import load_workbook
wb = load_workbook('sample.xlsx') # 파일의 wb를 불러옴 
ws = wb.active

for x in range(1, 11):
    for y in range(1, 11):
        print(ws.cell(row = x, column = y).value, end = " ")
    print()

# cell 갯수를 모르는 경우
maxRow    = ws.max_row
maxColumn = ws.max_column

for x in range(1, maxRow + 1):
    for y in range(1, maxColumn + 1):
        print(ws.cell(row = x, column = y).value, end = " ")
    print()

# wb.save('sample.xlsx') # 파일 저장
# wb.close() # 파일 종료 
