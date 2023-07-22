from openpyxl import Workbook
wb = Workbook() # 워크북 생성

ws = wb.active
ws.title = 'jonginSheet'

# 셀에 데이터 추가 
ws['A1'] = 1
ws['A2'] = 2
ws['A3'] = 3
ws['B1'] = 4
ws['C1'] = 5
ws['D1'] = 6

# 셀 데이터 획득 (없는 값은 none)
print(ws['A10'].value)

# Row는 1, 2, 3 Column은 A, B, C
# row = 1, column = 1 은 결국 A1과 같음 (ws['A1'])
print(ws.cell(row = 1, column = 1).value);

wb.save('sample.xlsx')
wb.close()